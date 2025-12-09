
# Standard Library Imports
import os          # Handles file paths and the file system.
import io          # For memory buffers (necessary for chart export function).
import itertools   # Used in the original code.

# External Library Imports
import pandas as pd     # Essential for reading and manipulating DataFrames.
import openpyxl         # For advanced Excel file manipulation.
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

# Visualization Libraries
import matplotlib.pyplot as plt 
import matplotlib.font_manager as font_manager

"==================================================================================================="

# Constants and Default Values
WATER_UNIT_WEIGHT = 9.81  # kN/m³ (Unit weight of water)
OUTPUT_FILENAME = "Results_Bearing_Capacity"
SHEET_NAME_1 = "Capacity_Charts"
SHEET_TITLE_1 = "Surface Bearing Capacity Results"
SHEET_NAME_2 = "settlement_check"
SHEET_TITLE_2 = "Settlement Check Results"
OUTPUT_DIR = "output"

"==================================================================================================="

def load_geotechnical_data(file_path: str = "Geotech_InputData.xlsx", 
                           sheet_name_1: str = "geotechnical_input", 
                           sheet_name_2: str = "footing_configuration") -> tuple:
    """
    Loads the main Excel file containing input data for shallow foundation design.
    
    Args:
        file_path: Path to the input Excel file (default: "Geotech_InputData.xlsx").
        sheet_name_1: Name of the first sheet (Geotechnical data) to load (default: "geotechnical_input").
        sheet_name_2: Name of the second sheet (Footing configuration) to load (default: "footing_configuration").
    
    Returns:
        A tuple containing (workbook, sheet_1, sheet_2) if loading is successful, 
        or (None, None, None) if an error occurs (e.g., File not found, missing sheet).
    
    Example:
        >>> workbook, geo_sheet, config_sheet = load_geotechnical_data()
        >>> if workbook:
        ...     # Process data...
        ...     pass
    """
    
    workbook = None
    sheet_1 = None
    sheet_2 = None

    try:
        workbook = openpyxl.load_workbook(file_path)

        # Load the first sheet (Geotechnical data)
        if sheet_name_1 in workbook.sheetnames:
            sheet_1 = workbook[sheet_name_1]
            print(f'✅ Sheet "{sheet_name_1}" loaded as sheet_1.')
        else:
            print(f'❌ Error: Sheet "{sheet_name_1}" not found in the file.')
            return None, None, None

        # Load the second sheet (Configuration/Results)
        if sheet_name_2 in workbook.sheetnames:
            sheet_2 = workbook[sheet_name_2]
            print(f'✅ Sheet "{sheet_name_2}" loaded as sheet_2.')
        else:
            print(f'❌ Error: Sheet "{sheet_name_2}" not found in the file.')
            return None, None, None

        print(f'✅ Excel file "{file_path}" loaded successfully.')
        return workbook, sheet_1, sheet_2

    except FileNotFoundError:
        print(f'❌ Error: File "{file_path}" not found. Check the path.')
        return None, None, None
    except Exception as e:
        print(f'⚠️ An error occurred while loading the file: {e}')
        return None, None, None
    
"==================================================================================================="

def read_row_vector(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                    row: int, 
                    col_start: int) -> list:
    """
    Reads a vector of values horizontally from an Excel sheet starting at a given cell (row, col_start).
    
    It iterates across columns until an empty cell is encountered.
    This is typically used to read footing dimensions (e.g., width B or embedment Df).
    
    Args:
        sheet: The openpyxl Worksheet object.
        row: The row number to start reading from (1-indexed).
        col_start: The starting column number (1-indexed).
        
    Returns:
        A list containing the values read from the row.
    """
    vector = []
    current_col = col_start
    while sheet.cell(row=row, column=current_col).value is not None:
        vector.append(sheet.cell(row=row, column=current_col).value)
        current_col += 1
    return vector

"==================================================================================================="

def read_column_vector(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                       row: int, 
                       col_start: int) -> list:
    """
    Reads a vector of geotechnical property values vertically from an Excel sheet 
    starting at a given cell (row, col_satart).
    
    It iterates down the column until an empty cell is encountered.
    This is typically used to read stratum properties (e.g., Cohesion 'c', Friction angle 'phi', 
    or Unit Weight 'gamma').
    
    Args:
        sheet: The openpyxl Worksheet object.
        row: The starting row number (1-indexed).
        col_start: The fixed column number (1-indexed).
        
    Returns:
        A list containing the values read from the column.
    """
    vector_column = []
    current_row = row
    while sheet.cell(row=current_row, column=col_start).value is not None:
        vector_column.append(sheet.cell(row=current_row, column=col_start).value)
        current_row += 1
    return vector_column

"==================================================================================================="

def export_dataframe_to_excel(results, output_dir, file_name, header_title):
    """
    Exports a DataFrame to an Excel file with custom formatting applied.

    The DataFrame rows are included without the original DataFrame headers.
    The column headers are placed in row 3 of the Excel sheet, and data starts in row 4.
    
    Args:
        results (pd.DataFrame): The DataFrame containing the results to export.
        output_dir (str): The destination folder path.
        file_name (str): The base name of the output Excel file (without extension).
        header_title (str): The main title for the spreadsheet (placed in cell A1).
    """
    # Imports are typically placed at the top of the file, 
    # but are kept here for self-contained execution if needed.
    # from openpyxl import load_workbook, ... 
    # import os, pandas as pd

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    excel_path = os.path.join(output_dir, f"{file_name}.xlsx")

    try:
        # Save to Excel, shifting data to start from row 4, without DataFrame header
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            results.to_excel(writer, index=False, startrow=3, header=False)

        # Load the Excel file to apply custom formatting
        wb = load_workbook(excel_path)
        ws = wb.active

        # Get the last column with data
        last_column = ws.max_column
        last_letter = get_column_letter(last_column)

        # Apply title formatting (Row 1)
        ws.merge_cells(f"A1:{last_letter}1")
        ws["A1"] = header_title
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(size=14, bold=True, name="Arial Narrow")
        ws["A1"].fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

        # Apply formatting to headers in Row 3
        for col_idx, col_name in enumerate(results.columns, start=1):
            column_letter = get_column_letter(col_idx)
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True, name="Arial Narrow")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.column_dimensions[column_letter].width = 17

        # Define borders for cells
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Apply borders and center alignment to all data cells
        last_row = ws.max_row
        for row in range(3, last_row + 1):
            for col in range(1, last_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Apply 2-decimal format to numeric values starting from Row 4
        for row in ws.iter_rows(min_row=4, max_row=last_row, min_col=1, max_col=last_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        # Save changes to the workbook
        wb.save(excel_path)
        wb.close()

        print(f"✅ DataFrame successfully exported to '{excel_path}' with formatting.")
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        
"==================================================================================================="

def export_charts_to_excel(figures_dict, output_dir, excel_filename):
    """
    Exports a dictionary of Matplotlib figures to a single Excel file,
    inserting them into the Excel file previously created by the data export function.

    Args:
        figures_dict (dict): A dictionary where keys are sheet names (str) 
                             and values are Matplotlib Figure objects (plt.Figure).
        output_dir (str): The destination folder path (e.g., 'output').
        excel_filename (str): The name of the Excel file (e.g., 'Results.xlsx').
    """
    # Construye la ruta completa: "output/NombreDelArchivo.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    try:
        # 1. Cargar el archivo Excel existente.
        # Esto fallará si la función de exportación de datos no ha creado el archivo primero.
        workbook_excel = load_workbook(excel_path)

        for sheet_name, figure in figures_dict.items():
            # 2. Obtener la hoja si existe, o crearla
            if sheet_name in workbook_excel.sheetnames:
                excel_sheet = workbook_excel[sheet_name]
            else:
                excel_sheet = workbook_excel.create_sheet(sheet_name) 

            # 3. Crear buffer en memoria para la imagen
            buffer = io.BytesIO()
            figure.savefig(buffer, format='png', bbox_inches='tight')
            buffer.seek(0)
            plt.close(figure) # Cierra la figura para liberar recursos
            
            # 4. Insertar la imagen en la hoja
            img = Image(buffer)
            excel_sheet.add_image(img, 'A1')

        # 5. Guardar el archivo una única vez al final, usando la ruta completa
        workbook_excel.save(excel_path)
        print(f"✅ Charts successfully exported to '{excel_path}'")

    except FileNotFoundError:
        # Mensaje de error clave para la depuración
        print(f"❌ Error: The file '{excel_path}' was not found. Ensure data export has created the file first.")
        return 
    except Exception as e:
        print(f"❌ Error exporting Charts: '{e}'")
        
"==================================================================================================="


def export_multiple_dataframes(df_1, df_2, output_dir, excel_filename, 
                               sheet_name_1, sheet_title_1, 
                               sheet_name_2, sheet_title_2):
    """
    Exports two DataFrames (df_1 and df_2) to two separate sheets in a single Excel file, 
    applying predefined custom formatting to both.

    Args:
        df_1 (pd.DataFrame): The first DataFrame to export (e.g., capacity results).
        df_2 (pd.DataFrame): The second DataFrame to export (e.g., settlement check results).
        output_dir (str): The destination folder path (e.g., 'output'). <--- ¡NUEVO!
        excel_filename (str): The base name of the Excel file (e.g., 'Results'). <--- ¡NUEVO!
        sheet_name_1 (str): Name for the first sheet. <--- ¡NUEVO!
        sheet_title_1 (str): Title for the first sheet. <--- ¡NUEVO!
        sheet_name_2 (str): Name for the second sheet. <--- ¡NUEVO!
        sheet_title_2 (str): Title for the second sheet. <--- ¡NUEVO!
    """
    
    # 1. Crear el directorio si no existe (usando el argumento output_dir)
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # 2. Construir la ruta completa
    excel_path = os.path.join(output_dir, f"{excel_filename}.xlsx")

    try:
        # Write both DataFrames to their respective sheets
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_1.to_excel(writer, sheet_name=sheet_name_1, startrow=3, index=False, header=False)
            df_2.to_excel(writer, sheet_name=sheet_name_2, startrow=3, index=False, header=False)

        # Cargar el libro de trabajo para aplicar el formato
        wb = load_workbook(excel_path)

        # Internal function to apply all formatting logic to a specific sheet
        def apply_format(ws, df, title, sheet_name): # Añadimos sheet_name como argumento
            # NOTA: ws = wb[sheet_name] ya está implícito en la lógica de abajo
            
            last_column = ws.max_column
            last_row = ws.max_row
            end_letter = get_column_letter(last_column)

            # Title (Row 1)
            ws.merge_cells(f"A1:{end_letter}1")
            ws["A1"] = title
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A1"].font = Font(size=14, bold=True, name="Arial Narrow")
            ws["A1"].fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")

            # Headers in Row 3
            for i, col_name in enumerate(df.columns, start=1):
                col_letter = get_column_letter(i)
                cell = ws.cell(row=3, column=i, value=col_name)
                cell.font = Font(bold=True, name="Arial Narrow")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                ws.column_dimensions[col_letter].width = 17

            # Borders definition
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )

            # Apply borders, alignment, and number format
            for row in range(3, last_row + 1):
                for col in range(1, last_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00"

        # Apply formatting to both sheets (using arguments instead of global variables)
        apply_format(wb[sheet_name_1], df_1, sheet_title_1, sheet_name_1)
        apply_format(wb[sheet_name_2], df_2, sheet_title_2, sheet_name_2)

        wb.save(excel_path)
        wb.close()
        print(f"✅ File successfully exported as '{excel_path}'.")

    except Exception as e:
        print(f"❌ Error exporting: {e}")